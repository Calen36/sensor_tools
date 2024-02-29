import csv
from datetime import date, datetime, timedelta
from os import path, walk
from statistics import mean
from tkinter import Tk

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, PatternFill

from data import combine_mean_datasets, get_lat_long, find_low_quality_monthly_values
from storage import ensure_dir_exists

import locale

locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')


def prepare_head(ws: Worksheet) -> None:
    """ Создает шапку для листа """
    col_widths = {'A': 13.85546875, 'B': 21, 'C': 13.85546875, 'D': 15.5703125, 'E': 13.140625, 'F': 18.5703125, 'G': 9.140625, 'H': 13.0}
    head = ['id устройства', 'Период', 'PM10, мкг/м³', 'PM2.5, мкг/м³', 'Влажность, %', 'Температура, t°C', 'lat', 'lon']

    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width
    ws.append(head)
    
    # Шрифт
    last_row = ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=False).__next__()
    for cell in last_row:
        cell.font = Font(name='Calibri', bold=True)


def apply_borders(ws: Worksheet) -> None:
    """ Задает границы всем ячейкам листа """
    border = Border(left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"))
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, values_only=False):
        for cell in row:
            cell.border = border


def form_period_repr(dt: datetime, period: str) -> str:
    """ Формирует содержимое ячейки Период в таблице """
    if period == 'hour':
        plus_hour = dt + timedelta(hours=1)
        return f"{dt.strftime('%Y.%m.%d')} {dt.strftime('%H:%M')}-{plus_hour.strftime('%H:%M')}"
    if period == 'day':
        return dt.strftime('%Y.%m.%d')
    if period == 'week':
        plus_6days = dt + timedelta(days=6)
        return f"{dt.strftime('%Y.%m.%d')} - {plus_6days.strftime('%Y.%m.%d')}"
    if period == 'month':
        # return f"{dt.strftime('%b').capitalize()} {dt.year}"
        return f"{dt.year}.{dt.month}"
    if period == 'year':
        return str(dt.year)
    return 'ERR'
    

def valdate_period(start_date: date, end_date: date, period: str, period_start: datetime) -> bool:
    """ Проверяем - полностью ли укладывается период period с началом в period_start в промежуток времени 
    между start_date и end_date (полностью включающий граничные дни)"""
    
    def get_last_second_of_month(dt: datetime) -> datetime:
        """ Возвращаем последнюю секунду месяца для заданного момента времени """
        next_month = dt.replace(day=28) + timedelta(days=4)
        last_day_of_month = next_month - timedelta(days=next_month.day)
        return last_day_of_month.replace(hour=23, minute=59, second=59)

    def get_last_second_of_year(dt: datetime) -> datetime:
        """ Возвращаем последнюю секунду года для заданного момента времени """
        year = dt.year + 1
        last_second = datetime(year, 1, 1) - timedelta(seconds=1)
        return last_second
    
    # данные периоды не требуют валидации
    if period.lower() in ('hour', 'day', 'week'):
        return True
    
    # формируем объекты datetime из date
    start_moment = datetime.combine(start_date, datetime.min.time())
    end_moment = datetime.combine(end_date, datetime.min.time()) + timedelta(days=1)
    # находим последнюю секунду периода
    if period.lower() == 'month':
        period_end = get_last_second_of_month(period_start)
    if period.lower() == 'year':
        period_end = get_last_second_of_year(period_start)

    if period_start < start_moment or period_end > end_moment:
        return False
    return True


def mark_low_qality_monthly_values(wb: Workbook,
                                   humid_data: dict[int: list[tuple[datetime, float, float]]],
                                   particle_data:  dict[int: list[tuple[datetime, float, float]]]):
    """ В excel таблице помечаем ячейки с малодостоверными месячными значениями красной заливкой.
    Малодостоверными считаем те месяцы, в которых данные по сенсорам есть менее чем за 15 дней из месяца."""
    red_fill = PatternFill(start_color='FF795C', end_color='FF795C', fill_type='solid')
    if 'Месяц' in wb.sheetnames:
        ws = wb['Месяц']
        # получаем список с единственным (в данном случае) списком месяцев с малодостоверными значениями
        bad_months_humid = list(find_low_quality_monthly_values(humid_data).values())
        bad_months_particle = list(find_low_quality_monthly_values(particle_data).values())
        # избавляемся от внешнего обертывающего списка
        bad_months_humid = bad_months_humid[0] if bad_months_humid else bad_months_humid
        bad_months_particle = bad_months_particle[0] if bad_months_particle else bad_months_particle
        # преобразуем объекты datetime в текстовые представления, которые используются в таблице
        bad_months_humid = [form_period_repr(dt, 'month') for dt in bad_months_humid]
        bad_months_particle = [form_period_repr(dt, 'month') for dt in bad_months_particle]

        # перебираем строки таблицы, если находим текущий месяц в списке плохих значений - отмечаем значение красной заливкой
        for row in ws:
            if len(row) > 3 and row[1].value in bad_months_particle:
                row[2].fill = red_fill
                row[3].fill = red_fill
            if len(row) > 5 and row[1].value in bad_months_humid:
                row[4].fill = red_fill
                row[5].fill = red_fill


def create_spreadsheet(workdir: str,
                       humid_data: dict[int: list[tuple[datetime, float, float]]],
                       particle_data: dict[int: list[tuple[datetime, float, float]]],
                       start_date: date,
                       end_date: date) -> str:
    """ СОЗДАНИЕ ТАБЛИЦЫ ИЗ ДАННЫХ 1го или 2х (разного типа) СЕНСОРОВ """

    sensor_ids = list(humid_data.keys()) + list(particle_data.keys())
    lat, lon = get_lat_long(workdir, sensor_ids[0], start_date, end_date)
    if len(sensor_ids) > 1:
        lat2, lon2 = get_lat_long(workdir, sensor_ids[1], start_date, end_date)
        if lat != lat2 or lon != lon2:
            print('ВНИМАНИЕ! Координаты заданных сеносоров не совпадают!')

    """ создаем объединенный словарь, ключи - наименование периодов усредения (day, month, и т.д.).
    Значения - также словарь, где ключами являются объекты datetime начала периода.
    Значения 2го порядка - словарь с содержимым строки. Ключи тут: 'ids', 'p1', 'p2', 'humid', 'temp'. """
    combined_data = combine_mean_datasets(humid_data=humid_data, particle_data=particle_data)

    wb = Workbook()
    wb.remove(wb.active)
    
    periods = (('hour', 'Час'), ('day', 'Сутки'), 
               ('week', 'Неделя'), ('month', 'Месяц'), ('year', 'Год'))

    for period, sheet_name in periods:  # ИТЕРАЦИЯ ПО ПЕРИОДАМ УСРЕДЕНИЯ
        sheet = wb.create_sheet(sheet_name)  # создаем лист для каждого периода
        prepare_head(sheet)  # создание шапки листа
        for dt in sorted(combined_data[period].keys()):  # ИТЕРАЦИЯ ПО ВРЕМЕНАМ НАЧАЛА ПЕРИОДА
            # добавление строки
            if valdate_period(start_date, end_date, period, dt):
                sheet.append([', '.join([str(id_) for id_ in sorted(combined_data[period][dt]['ids'])]),
                            form_period_repr(dt, period),
                            combined_data[period][dt]['p1'],
                            combined_data[period][dt]['p2'],
                            combined_data[period][dt]['humid'],
                            combined_data[period][dt]['temp'],
                            lat,
                            lon])
        apply_borders(sheet)  # задаение границ всем ячейкам листа

        # если в лист в итоге пуст - удаляем его
        if sheet.max_row <= 1:
            wb.remove(sheet)

    # помечаем малодостоверные месячные значения.
    mark_low_qality_monthly_values(wb, humid_data, particle_data)

    # СОХРАНЕНИЕ ФАЙЛА        
    ids = list(humid_data.keys()) + list(particle_data.keys())
    ids_repr = ', '.join([str(id_) for id_ in sorted(ids)])
    out_dir = path.normpath(path.join(workdir, 'OUTPUT'))
    ensure_dir_exists(out_dir)
    out_filename = path.normpath(path.join(out_dir, f'{ids_repr}_{start_date}_{end_date}.xlsx'))
    print(f'Сохранен файл {out_filename}')    
    wb.save(out_filename)
    return out_filename


if __name__ == "__main__":
    pass
